"""Oracle: every dashboard model figure names its source cells.

The trust claim is literal: a number shown to a client traces to exact
cells of the uploaded workbook. Series extracted from matrix and tidy
tables carry per-period source-cell refs (A1, absolute even when the
table sat in a side panel); model metrics, plan-vs-actual and the
unit-cost-vs-budget target expose them; and re-reading every emitted ref
from the original bytes must reproduce the rendered value. A figure whose
cells can't be named simply carries no ``cells`` key — never a guess.
"""
from __future__ import annotations

import io
import re

import openpyxl
import pandas as pd
from fastapi.testclient import TestClient

from app.main import app
from app.snapshot import build_exec_snapshot

client = TestClient(app)

MONTHS = ["Juil 2025", "Août 2025", "Sept 2025", "Oct 2025",
          "Nov 2025", "Déc 2025"]
CPO = [10, 40, 50, 100, 80, 16]
PLAN_CPO = [20, 80, 100, 200, 160, 32]
OPEX = [10, 10, 10, 10, 10, 10]
PLAN_OPEX = [20, 20, 20, 20, 20, 20]


def _book(*sheets: tuple) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for name, df in sheets:
            df.to_excel(xw, sheet_name=name, header=False, index=False)
    return buf.getvalue()


def _analyze(content: bytes, name="pvak.xlsx") -> dict:
    r = client.post("/analyze", files={
        "file": (name, content, "application/octet-stream")})
    assert r.status_code == 200, r.text
    return r.json()


def _sheet(banner: str, rows: list, months=MONTHS) -> pd.DataFrame:
    # Excel shape: banner row 1, header row 2, data rows from row 3.
    return pd.DataFrame([[banner, *([None] * len(months))],
                         ["SERIE", *months], *rows])


def _actuals() -> pd.DataFrame:
    return _sheet("JOURNAL PRODUCTION", [
        ["PRODUCTION CPO (T)", *CPO],          # row 3, values B3..G3
        ["CHARGES OPEX", *OPEX],               # row 4, values B4..G4
    ])


def _plan() -> pd.DataFrame:
    return _sheet("PLAN OPERATIONNEL", [
        ["PRODUCTION CPO PREVUE (T)", *PLAN_CPO],   # row 3
        ["OPEX PREVISIONNEL", *PLAN_OPEX],          # row 4
    ])


def _row_refs(row: int, n: int = 6) -> list:
    return [[f"{chr(ord('B') + k)}{row}"] for k in range(n)]


def test_matrix_series_carry_cells():
    rep = _analyze(_book(("JOURNAL", _actuals()), ("BUDGET 2025", _plan())))
    met = rep["model"]["monthly"]["metrics"]
    assert met["volume"]["cells"] == _row_refs(3)
    assert met["opex"]["cells"] == _row_refs(4)
    pva = rep["model"]["monthly"]["plan_vs_actual"]
    assert pva["volume"]["plan_cells"] == _row_refs(3)
    assert pva["opex"]["plan_cells"] == _row_refs(4)


def test_duplicate_label_rows_list_every_contributing_cell():
    # Two rows share one label: the series value is their sum, and BOTH
    # cells are named — a partial citation would be a fabrication.
    df = _sheet("JOURNAL PRODUCTION", [
        ["PRODUCTION CPO (T)", *CPO],          # row 3
        ["PRODUCTION CPO (T)", *CPO],          # row 4 — same label
    ])
    rep = _analyze(_book(("JOURNAL", df)))
    vol = rep["model"]["monthly"]["metrics"]["volume"]
    assert vol["values"] == [v * 2 for v in CPO]
    assert vol["cells"] == [[f"{chr(ord('B') + k)}3", f"{chr(ord('B') + k)}4"]
                            for k in range(6)]


def test_panel_cells_are_sheet_absolute():
    # Two side-by-side matrices split into panels; the right panel's refs
    # must be absolute sheet addresses, not slice-relative ones.
    left = _sheet("NOTES", [["DIVERS", *([1] * 6)]])
    right = _sheet("JOURNAL", [["PRODUCTION CPO (T)", *CPO]])
    combined = pd.concat(
        [left, pd.DataFrame([[None]] * len(left)), right], axis=1)
    rep = _analyze(_book(("FEUILLE", combined)))
    sheet = rep["sheets"][0]
    assert sheet["orientation"] == "multi", sheet["orientation_reason"]
    chart = sheet["panels"][1]["tidy"]["summary"]["chart"]
    ser = {s["label"]: s for s in chart["series_all"]}
    cells = ser["PRODUCTION CPO (T)"]["cells"]
    # right panel starts after 7 left columns + 1 blank: sheet column J on
    assert cells == [[f"{chr(ord('J') + k)}3"] for k in range(6)]


def _unit_cost_book() -> bytes:
    """The anchor shape from test_month_axis, budget sheet included."""
    rows = [
        ["Actuals — Q1 2026", None, None, None, None],
        [None, None, None, None, None],
        ["Metric", "Jan 2026", "Feb 2026", "Mar 2026", "Q1 Total"],
        ["PRODUCTION", None, None, None, None],
        ["CPO produced (t)", 16, 17.5, 18.5, 52],       # row 5
        ["FFB produced (t)", 76, 80, 84, 240],          # row 6
        ["CASH OUT", None, None, None, None],
        ["OPEX — production cost (USD)", 14200, 16900, 15400, 46500],  # row 8
        ["CAPEX — investment (USD)", 41000, 52500, 49800, 143300],     # row 9
    ]
    budget = [
        ["Metric", 2025, 2026, 2027],
        ["FFB produced (t)", 3000, 3200, 4500],          # row 2
        ["CPO produced (t)", 650, 700, 990],             # row 3
        ["OPEX — production cost (USD)", 351000, 378000, 534600],  # row 4
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        pd.DataFrame(rows).to_excel(w, sheet_name="Actuals", header=False,
                                    index=False)
        pd.DataFrame(budget).to_excel(w, sheet_name="BUDGET 2025-2027",
                                      header=False, index=False)
    return buf.getvalue()


def test_unit_cost_budget_names_its_target_cells():
    rep = _analyze(_unit_cost_book(), name="model.xlsx")
    mo = rep["model"]["monthly"]
    # the monthly actual side already traces through metrics[...].cells
    assert mo["metrics"]["opex"]["cells"] == [["B8"], ["C8"], ["D8"]]
    assert mo["metrics"]["volume_secondary"]["cells"] == \
        [["B5"], ["C5"], ["D5"]]
    src = mo["unit_cost_budget"]["target_sources"]
    assert src["opex"] == {"sheet": "BUDGET 2025-2027",
                           "label": "OPEX — production cost (USD)",
                           "cells": ["C4"]}
    assert src["volume_secondary"] == {"sheet": "BUDGET 2025-2027",
                                       "label": "CPO produced (t)",
                                       "cells": ["C3"]}


def test_tidy_year_rows_carry_cells():
    # A tidy table with a year column: the year series live in ROWS, so a
    # per-year value cites the row(s) it was summed from.
    df = pd.DataFrame([
        ["Année", "PRODUCTION CPO (T)", "CHARGES OPEX"],
        [2023, 500, 90],       # row 2
        [2024, 650, 100],      # row 3
        [2025, 700, 120],      # row 4
    ])
    rep = _analyze(_book(("PROJECTIONS", df)))
    chart = rep["sheets"][0]["tidy"]["summary"]["chart"]
    assert chart["axis"] == "year"
    ser = {s["label"]: s for s in chart["series_all"]}
    assert ser["PRODUCTION CPO (T)"]["cells"] == [["B2"], ["B3"], ["B4"]]
    assert ser["CHARGES OPEX"]["cells"] == [["C2"], ["C3"], ["C4"]]


_A1 = re.compile(r"^([A-Z]+)(\d+)$")


def _read_ref(ws, ref: str):
    m = _A1.match(ref)
    assert m, f"malformed cell ref {ref!r}"
    return ws[ref].value


def _assert_series_reproduces(ws, values, cells):
    assert len(cells) == len(values)
    for v, refs in zip(values, cells):
        if v is None:
            continue
        assert refs, f"value {v} carries no source cells"
        total = sum(float(_read_ref(ws, r)) for r in refs)
        assert round(total, 4) == v, (refs, total, v)


def test_every_emitted_ref_reproduces_the_value():
    # The anti-fabrication property: open the uploaded bytes at every ref
    # the model emitted and re-derive the number shown.
    content = _unit_cost_book()
    rep = _analyze(content, name="model.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    mo = rep["model"]["monthly"]
    for role, m in mo["metrics"].items():
        ws = wb[m["sheet"]]
        _assert_series_reproduces(ws, m["values"], m["cells"])
    for role, m in rep["model"]["metrics"].items():
        ws = wb[m["sheet"]]
        _assert_series_reproduces(ws, m["values"], m["cells"])


def test_plan_cells_reproduce_the_plan():
    content = _book(("JOURNAL", _actuals()), ("BUDGET 2025", _plan()))
    rep = _analyze(content)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    for role, p in rep["model"]["monthly"]["plan_vs_actual"].items():
        ws = wb[p["plan_sheet"]]
        _assert_series_reproduces(ws, p["plan"], p["plan_cells"])


def test_band_offset_cells_are_sheet_absolute():
    # A matrix classified inside a stacked BAND (rows offset) must cite
    # absolute sheet rows, exactly like the panel (column) case.
    from tests.fixtures import stacked_tables_sheet
    from app.pipeline.orient import orient_sheet
    out = orient_sheet(stacked_tables_sheet())
    assert out["orientation"] == "multi"
    matrix = next(p for p in out["panels"]
                  if (p.get("tidy") or {}).get("columns", [None])[-1] == "value")
    ser = {s["label"]: s
           for s in matrix["tidy"]["summary"]["chart"]["series_all"]}
    # FFB RECU lives at sheet row 19, CPO PRODUIT at row 20 (band offset 17)
    assert ser["FFB RECU"]["cells"][0] == ["B19"]
    assert ser["CPO PRODUIT"]["cells"][0] == ["B20"]


def test_total_rows_never_enter_a_roles_cells():
    # The TOTAL series is an aggregate, not data: role metrics cite only the
    # real rows, and no citation anywhere names the TOTAL row.
    df = _sheet("JOURNAL PRODUCTION", [
        ["PRODUCTION CPO (T)", *CPO],           # row 3
        ["CHARGES OPEX", *OPEX],                # row 4
        ["TOTAL PRODUCTION", *[c + o for c, o in zip(CPO, OPEX)]],   # row 5
    ])
    rep = _analyze(_book(("JOURNAL", df)))
    met = rep["model"]["monthly"]["metrics"]
    assert met["volume"]["cells"] == _row_refs(3)
    for role, m in met.items():
        for refs in m["cells"]:
            assert not any(r.endswith("5") for r in refs or []), (role, refs)


def test_short_plan_cells_pad_with_none_like_the_values():
    # A 3-month plan against 6 months of actuals: plan_cells align to the
    # actuals axis exactly as plan values do — None where the plan is silent.
    plan3 = _sheet("PLAN OPERATIONNEL", [
        ["PRODUCTION CPO PREVUE (T)", *PLAN_CPO[:3]],
    ], MONTHS[:3])
    rep = _analyze(_book(("JOURNAL", _actuals()), ("BUDGET 2025", plan3)))
    v = rep["model"]["monthly"]["plan_vs_actual"]["volume"]
    assert v["plan"] == [*PLAN_CPO[:3], None, None, None]
    assert v["plan_cells"] == [*_row_refs(3, 3), None, None, None]


def test_no_positions_means_no_cells_key():
    # Provenance is never guessed: a chart built without positional info
    # simply carries no ``cells`` key at all.
    from app.pipeline.charts import year_series_chart
    ch = year_series_chart(
        {"Année": [2023, 2024, 2025], "PRODUCTION CPO (T)": [500, 650, 700]},
        set())
    assert ch["axis"] == "year"
    for s in ch["series_all"]:
        assert "cells" not in s


def test_ffilled_label_columns_never_become_numeric_series():
    # The value-side ghost fix: a label-ish column (majority text among its
    # RAW cells) that gets forward-filled must never profile numeric and
    # become a row-year series — its sums would count merged-cell ghosts
    # (one real 500 rendering as 1500). Label columns label; they are not
    # data series.
    df = pd.DataFrame([
        ["Année", "PRODUCTION CPO (T)", "CHARGES OPEX", "SURFACE (HA)"],
        [2023, 500, 90, 12],           # the ONLY real value in column B
        [2023, None, 100, 12],
        [2023, None, 95, 13],
        [2024, None, 110, 13],
        [2024, None, 105, 14],
        [2025, "voir note", 120, 14],
        [2025, "voir note", 115, 15],
    ])
    rep = _analyze(_book(("PROJECTIONS", df)))
    chart = rep["sheets"][0]["tidy"]["summary"]["chart"]
    labels = {s["label"] for s in chart["series_all"]}
    assert "PRODUCTION CPO (T)" not in labels, labels
    # and no ghost-inflated figure leaks into the model anywhere
    for role, m in ((rep.get("model") or {}).get("metrics") or {}).items():
        assert 1500.0 not in (m["values"] or []), (role, m)
    # the honest series still sum their real cells
    ser = {s["label"]: s for s in chart["series_all"]}
    assert ser["CHARGES OPEX"]["values"] == [285, 215, 235]


def test_ffilled_ghost_rows_block_the_citation():
    # A value forward-filled from a merged cell sits on a physically BLANK
    # cell. Citing it — or only the real subset of a summed period — would
    # be a false proof, so such periods carry no refs at all.
    from app.pipeline.charts import year_series_chart
    ch = year_series_chart(
        {"Année": [2023, 2023, 2024, 2025],
         "PRODUCTION CPO (T)": [500, 500, 650, 700]},
        set(),
        row_numbers=[2, 3, 4, 5],
        col_map={"PRODUCTION CPO (T)": 1},
        ghost_rows={"PRODUCTION CPO (T)": {1}})    # ordinal 1 ffilled from B2
    ser = {s["label"]: s for s in ch["series_all"]}
    cells = ser["PRODUCTION CPO (T)"]["cells"]
    assert cells[0] is None                        # 2023 summed a ghost
    assert cells[1] == [[4, 1]] and cells[2] == [[5, 1]]   # real rows cite


def test_public_snapshot_carries_no_cells():
    # The published exec payload carries no file structure — cell refs
    # included, same policy as sheet names.
    rep = _analyze(_unit_cost_book(), name="model.xlsx")
    snap = build_exec_snapshot(rep)

    def walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                assert k not in ("cells", "plan_cells", "target_sources"), k
                walk(v)
        elif isinstance(node, list):
            for v in node:
                walk(v)

    walk(snap["model"])
