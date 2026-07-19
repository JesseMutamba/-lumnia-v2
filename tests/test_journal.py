"""The journal engine: contract-based audit of dual cash journals.

The fixture replicates the SUIVI FINANCES shape — header at sheet row 10,
dual-currency mirroring, a DGO journal re-keyed into the consolidated one,
an EXTRACTION MOIS glossary with section headers — with ONE planted
violation per rule. The tie-out must hold to the franc because unmapped
codes land in an explicit UNMAPPED bucket, never vanish.
"""
from __future__ import annotations

import io

import openpyxl
from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)

PIVOT = 2250


def _journal_sheet(ws, headers, rows):
    ws["A1"] = "JOURNAL DE CAISSE"          # banner rows 1-9, header row 10
    for c, h in enumerate(headers, start=1):
        ws.cell(row=10, column=c, value=h)
    for ri, row in enumerate(rows, start=11):
        for c, v in enumerate(row, start=1):
            ws.cell(row=ri, column=c, value=v)


def _book() -> bytes:
    wb = openpyxl.Workbook()
    site = wb.active
    site.title = "JC Consolidé"
    h = ["Dates", "Référence Liée", "Personne/Service", "Libellés",
         "Code Interne", "CODE", "Entrées CDF", "Sorties CDF", "Solde CDF",
         "Entrées USD", "Sorties USD", "Solde USD"]
    # balances tracked so V6 fires exactly once (the planted break)
    site_rows = [
        # opening float
        ["2026-01-02", "R1", "Caisse", "Report de solde", "REPORT", "571100",
         10_000_000, 0, 10_000_000, 5000, 0, 5000],
        # plain CDF spend
        ["2026-01-05", "R2", "RH", "Salaires janvier", "SAL001", "661100",
         0, 4_500_000, 5_500_000, 0, 0, 5000],
        # dual-currency mirror: USD spend echoed in CDF (spend = CDF sortie)
        ["2026-01-10", "R3", "Log", "Carburant groupe", "FUEL01", "602200",
         450_000, 450_000, 5_500_000, 0, 200, 4800],
        # the site echo of the matched DGO outflow (500 USD × 2250)
        ["2026-01-15", "R4", "DGO", "Réplication DGO intrants", "AGRI01",
         "602100", 0, 1_125_000, 4_375_000, 0, 0, 4800],
        # V3: bank transfer recorded as spend under a neutral code
        ["2026-02-04", "R5", "Rawbank", "Dépôt Rawbank", "Banque", "521100",
         0, 1_800_000, 2_575_000, 0, 0, 4800],
        # V2: libellé present, all amounts zero
        ["2026-02-10", "R6", "DGDA", "Taxes DGDA à régulariser", "TAX001",
         "641100", 0, 0, 2_575_000, 0, 0, 4800],
        # V6: balance break — solde jumps without a movement explaining it
        ["2026-02-15", "R7", "Log", "Entretien pistes", "OPX002", "615200",
         0, 500_000, 1_575_000, 0, 0, 4800],   # should be 2_075_000
        # V4 + UNMAPPED: code absent from the glossary
        ["2026-03-03", "R8", "RH", "Prime équipe usine", "RH", "661200",
         0, 2_250_000, -675_000, 0, 0, 4800],
        # V5: out of date order (March row dated February)
        ["2026-02-20", "R9", "Log", "Achat pièces", "OPX002", "615200",
         0, 900_000, -1_575_000, 0, 0, 4800],
        # V8: exact duplicate of the pièces row
        ["2026-02-20", "R9", "Log", "Achat pièces", "OPX002", "615200",
         0, 900_000, -2_475_000, 0, 0, 4800],
    ]
    _journal_sheet(site, h, site_rows)

    dgo = wb.create_sheet("JC DGO")
    hd = ["Dates", "CPT", "N°PCS", "Libellés", "Code Interne", "CODE",
          "Entrées USD", "Sorties USD", "Solde USD"]
    dgo_rows = [
        ["2026-01-03", "DG", "P1", "Dotation janvier", "REPORT", "571100",
         3000, 0, 3000],
        # matched by the site échon (500 USD ≈ 1,125,000 CDF)
        ["2026-01-15", "DG", "P2", "Intrants agricoles", "AGRI01", "602100",
         0, 500, 2500],
        # V1: orphan — no site row anywhere near 1,200 USD in February
        ["2026-02-18", "DG", "P3", "Replanting pépinière", "CAPX01", "231100",
         0, 1200, 1300],
    ]
    _journal_sheet(dgo, hd, dgo_rows)

    ex = wb.create_sheet("EXTRACTION MOIS")
    ex["E4"] = PIVOT
    layout = [
        ("OPEX — Administration", None),
        ("Salaires", "SAL001"), ("Taxes", "TAX001"),
        ("OPEX — Plantations", None),
        ("Intrants", "AGRI01"), ("Entretien", "OPX002"),
        ("OPEX — Mill (Techniques)", None),
        ("Carburant", "FUEL01"),
        ("CAPEX — Replanting", None),
        ("Pépinière", "CAPX01"),
        ("Head Office (DGO)", None),
        ("Report", "REPORT"), ("Banque", "Banque"),
    ]
    for ri, (label, code) in enumerate(layout, start=6):
        ex.cell(row=ri, column=2, value=label)
        if code:
            ex.cell(row=ri, column=3, value=code)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _journal() -> dict:
    r = client.post("/analyze", files={
        "file": ("SUIVI FINANCES PVAK V1.xlsx", _book(),
                 "application/octet-stream")})
    assert r.status_code == 200, r.text
    j = r.json()["journal"]
    assert j is not None
    return j


def test_every_rule_fires_exactly_on_the_planted_violation():
    j = _journal()
    by_rule = {}
    for f in j["findings"]:
        by_rule.setdefault(f["rule"], []).append(f)
    assert len(by_rule.get("V1", [])) == 1        # the 1,200 USD orphan
    assert by_rule["V1"][0]["usd"] == 1200
    assert len(by_rule.get("V2", [])) == 1        # zero-amount taxes row
    assert by_rule["V2"][0]["code"] == "TAX001"
    assert len(by_rule.get("V3", [])) == 1        # Rawbank deposit as spend
    assert by_rule["V3"][0]["usd"] == 800         # 1.8M CDF / 2250
    assert [f["code"] for f in by_rule.get("V4", [])] == ["RH"]
    assert by_rule["V4"][0]["usd"] == 1000        # 2.25M CDF / 2250
    assert len(by_rule.get("V5", [])) >= 1        # out-of-order pièces row
    assert len(by_rule.get("V6", [])) == 1        # the planted solde break
    assert len(by_rule.get("V8", [])) == 1        # duplicated pièces row
    # findings carry both languages
    assert all(f["msg"] and f["msg_fr"] for f in j["findings"])


def test_tie_out_holds_to_the_franc_via_the_unmapped_bucket():
    j = _journal()
    assert not any(f["rule"] == "V7" for f in j["findings"])
    total_cats = round(sum(sum(c["cdf"]) for c in j["cats"]), 2)
    assert total_cats == j["meta"]["tie_out_cdf"]
    # site journal: 4.5M + 450k + 1.125M + 1.8M + 500k + 2.25M + 2×900k
    assert j["meta"]["tie_out_cdf"] == 12_425_000
    un = next(c for c in j["cats"] if c["code"] == "UNMAPPED")
    assert sum(un["cdf"]) == 2_250_000            # the RH prime, visible


def test_aggregation_is_code_by_month_with_dual_currency_convention():
    j = _journal()
    assert j["months"] == ["2026-01", "2026-02", "2026-03"]
    fuel = next(c for c in j["cats"] if c["code"] == "FUEL01")
    # the USD-paid fuel is counted ONCE via its CDF sortie mirror
    assert fuel["cdf"] == [450_000, 0, 0]
    assert fuel["usd"] == [200, 0, 0]
    sal = next(c for c in j["cats"] if c["code"] == "SAL001")
    assert sal["sec"] == "OPEX — Administration"
    assert sal["kind"] == "opex"
    capx = next(c for c in j["cats"] if c["code"] == "CAPX01")
    assert capx["kind"] == "capex"
    assert j["meta"]["rate"] == PIVOT
    assert j["meta"]["dgo_out_usd"] == 1700       # 500 matched + 1200 orphan


def test_non_contract_workbooks_get_no_journal_block():
    import pandas as pd
    import io as _io
    df = pd.DataFrame([["A", "B"], [1, 2]])
    buf = _io.BytesIO()
    df.to_excel(buf, header=False, index=False, engine="openpyxl")
    r = client.post("/analyze", files={
        "file": ("plain.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.json()["journal"] is None
